import abc
from asyncio.log import logger
import csv
import os
import pathlib
from openpyxl import load_workbook
import xlrd
import math
import shutil
import traceback
from datetime import datetime
from zipfile import ZipFile


def fatal_error(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as ex:
            logger.warning(traceback.format_exc())
            exit()

    return wrapper


def warning_error(func):
    def wrapper(*args):
        try:
            return func(*args)
        except Exception as ex:
            template = "An exception of type {0} occurred. Arguments:\n{1!r}"
            message = template.format(type(ex).__name__, ex.args)
            logger.error(message)
            return None

    return wrapper


def rchop(s, sub):
    return s[: -len(sub)] if s.endswith(sub) else s


class DataFile(abc.ABC):
    def __init__(
        self, fname, sheet_name: str = "", first_line: int = 0, columns: range = [50]
    ):
        self._fname = fname
        self._first_line = first_line
        self._sheet_name = sheet_name
        self._columns = columns

    def __iter__(self):
        return self

    def __next__(self):
        return ""


class XlsFile(DataFile):
    def __init__(
        self,
        fname,
        sheet_name: str = "",
        first_line: int = 0,
        columns: int = 50,
        page_index=0,
    ):
        super(XlsFile, self).__init__(fname, sheet_name, first_line, range(columns))
        self._book = xlrd.open_workbook(
            fname, logfile=open(os.devnull, "w"), ignore_workbook_corruption=True
        )
        self.set_current_sheet()

    def set_current_sheet(self, sheet_name: str = "", page_index: int = 0):
        if sheet_name:
            sheet = self._book.sheet_by_name(self._sheet_name)
        else:
            sheet = self._book.sheets()[page_index]
        self._rows = (self.sheet.row(index) for index in range(sheet.nrows))

    @staticmethod
    def get_cell_text(cell):
        if cell.ctype == 2:
            return rchop(str(cell.value), ".0")
        return str(cell.value)

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self._columns:
                yield XlsFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._rows:
            return list(self.get_row(row))
        raise StopIteration

    def __del__(self):
        pass


class XlsxFile(DataFile):
    @fatal_error
    def __init__(
        self,
        fname,
        sheet_name: str = "",
        first_line: int = 0,
        columns: int = 50,
        page_index: int = 0,
    ):
        super(XlsxFile, self).__init__(fname, sheet_name, first_line, range(columns))

        self._wb = load_workbook(filename=fname, read_only=True, data_only=True)
        self.set_current_sheet()
    
    def set_current_sheet(self, sheet_name: str = "", page_index: int = 0):
        if sheet_name:
            self._ws = self._wb.get_sheet_by_name(sheet_name)
        else:
            self._ws = self._wb.worksheets[page_index]
        self._cursor = self._ws.iter_rows()
        row_num = 0
        while row_num < self._first_line:
            row_num += 1
            next(self._cursor)
    
    @staticmethod
    def get_cell_text(cell):
        return str(cell.value) if cell.value else ""

    def get_row(self, row):
        i = 0
        for cell in row:
            if i in self._columns:
                yield XlsxFile.get_cell_text(cell)
            i += 1

    def __next__(self):
        return list(self.get_row(next(self._cursor)))

    def __del__(self):
        self._wb.close()

    def get_index(self, cell):
        try:
            return cell.column
        except AttributeError:
            return -1


def get_file_reader(fname):
    _, file_extension = os.path.splitext(fname)
    if file_extension == ".xls":
        return XlsFile
    if file_extension == ".xlsx":
        return XlsxFile
    raise Exception("Unknown file type")
